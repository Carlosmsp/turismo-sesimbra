"""
rotas.py — define todas as rotas HTTP da aplicação Flask:
  API pública (/api/chatbot, /api/conteudo, /api/recomendacoes, /contactos, /api/sugestoes)
  API do painel admin (/api/admin/...) — protegida por token
  Ficheiros estáticos (HTML, assets, imagens)
"""
from __future__ import annotations

import io
import json
import secrets
import sqlite3
from datetime import datetime
from types import SimpleNamespace

from flask import Response, jsonify, request, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


CAMPOS_EDITAVEIS_CONTEUDO = {
    "pontos_interesse": ["nome", "descricao", "local", "imagem", "telefone", "site_url", "lat", "lon", "aviso", "recomendado_por", "ordem"],
    "restaurantes": ["nome", "descricao", "tipo", "local", "imagem", "telefone", "email", "site_url", "lat", "lon", "recomendado_por", "ordem"],
    "alojamentos": ["nome", "descricao", "local", "imagem", "telefone", "site_url", "booking_url", "lat", "lon", "recomendado_por", "ordem"],
    "atividades": ["nome", "descricao", "local", "telefone", "site_url", "links_json", "icone", "categoria", "lat", "lon", "recomendado_por", "ordem"],
}


def registar_rotas(app, deps: SimpleNamespace) -> None:
    @app.post("/api/chatbot")
    def api_chatbot():
        # recebe a pergunta do utilizador e devolve a resposta do assistente ia
        dados = request.get_json(silent=True) or {}
        pergunta = str(dados.get("pergunta", "")).strip()
        pagina = str(dados.get("pagina", "sesimbra.html")).strip() or "sesimbra.html"
        historico_recebido = dados.get("historico")
        historico = (
            [item for item in historico_recebido if isinstance(item, dict)]
            if isinstance(historico_recebido, list)
            else []
        )
        bloco_historico = deps.formatar_historico_chatbot(historico)

        if not pergunta:
            return jsonify({"erro": "A pergunta e obrigatoria."}), 400

        ip_cliente = deps.obter_ip_cliente()
        permitido, mensagem_limite = deps.verificar_limite_ip(ip_cliente)
        if not permitido:
            resposta_site = {
                "modelo": deps.GEMINI_MODEL,
                "estado": "limitado",
                "resposta": mensagem_limite,
            }
            deps.guardar_log_api(
                endpoint="/api/chatbot",
                modelo=deps.GEMINI_MODEL,
                pedido={**dados, "ip": ip_cliente},
                resposta={"gemini": None, "resposta_site": resposta_site},
                prompt="Pedido bloqueado por limite de utilização por IP.",
                resposta_text=mensagem_limite,
                estado="limitado",
            )
            return jsonify(resposta_site), 429

        meteorologia = deps.obter_meteorologia()
        transporte_tempo_real = deps.obter_transporte_tempo_real(pergunta)
        prompt = deps.criar_prompt_chatbot(pergunta, pagina, meteorologia, transporte_tempo_real, bloco_historico)
        texto, resposta_gemini, estado = deps.chamar_gemini_prompt(prompt)

        resposta_site = {
            "modelo": deps.GEMINI_MODEL,
            "estado": estado,
            "resposta": texto,
        }

        deps.guardar_log_api(
            endpoint="/api/chatbot",
            modelo=deps.GEMINI_MODEL,
            pedido=dados,
            resposta={"gemini": resposta_gemini, "resposta_site": resposta_site},
            prompt=prompt,
            resposta_text=texto,
            estado=estado,
        )

        return jsonify(resposta_site)

    @app.route("/api/chatbot", methods=["OPTIONS"])
    def api_chatbot_options():
        return "", 204

    @app.post("/contactos")
    @app.post("/api/contactos")
    def api_contactos():
        # valida csrf, aplica rate limit e guarda a mensagem de contacto na base de dados
        if not deps.validar_csrf():
            return jsonify({"estado": "erro", "erros": ["Pedido inválido. Atualize a página e tente novamente."]}), 403

        permitido, mensagem_limite = deps.verificar_limite_contactos(deps.obter_ip_cliente())
        if not permitido:
            return jsonify({"estado": "erro", "erros": [mensagem_limite]}), 429

        dados = request.get_json(silent=True) or request.form.to_dict()
        contacto, erros = deps.validar_dados_contacto(dados)

        if erros:
            return jsonify({"estado": "erro", "erros": erros}), 400

        criado_em = deps.guardar_contacto(
            nome=contacto["nome"],
            email=contacto["email"],
            telefone=contacto["telefone"],
            mensagem=contacto["mensagem"],
        )

        return jsonify(
            {
                "estado": "ok",
                "mensagem": "Mensagem guardada com sucesso.",
                "criado_em": criado_em,
            }
        ), 201

    @app.route("/contactos", methods=["OPTIONS"])
    @app.route("/api/contactos", methods=["OPTIONS"])
    def api_contactos_options():
        return "", 204

    @app.post("/api/sugestoes")
    def api_sugestoes():
        # valida csrf, aplica rate limit, guarda a foto (se houver) e a sugestão
        if not deps.validar_csrf():
            return jsonify({"estado": "erro", "erros": ["Pedido inválido. Atualize a página e tente novamente."]}), 403

        permitido, mensagem_limite = deps.verificar_limite_contactos(deps.obter_ip_cliente())
        if not permitido:
            return jsonify({"estado": "erro", "erros": [mensagem_limite]}), 429

        dados = request.form.to_dict() if request.form else (request.get_json(silent=True) or {})
        foto_path = ""
        foto = request.files.get("foto")
        if foto and foto.filename:
            foto_path, erro_foto = deps.guardar_imagem_sugestao(foto)
            if erro_foto:
                return jsonify({"estado": "erro", "erros": [erro_foto]}), 400

        dados["foto_path"] = foto_path
        sugestao, erros = deps.validar_sugestao_local(dados)

        if erros:
            # apaga a foto já guardada se a validação falhar
            caminho_foto = deps.BASE_DIR / foto_path
            if caminho_foto.exists() and caminho_foto.is_file():
                caminho_foto.unlink()
            return jsonify({"estado": "erro", "erros": erros}), 400

        _, criado_em = deps.guardar_sugestao_local(sugestao)

        return jsonify(
            {
                "estado": "ok",
                "mensagem": "Sugestão guardada com sucesso.",
                "criado_em": criado_em,
            }
        ), 201

    @app.route("/api/sugestoes", methods=["OPTIONS"])
    def api_sugestoes_options():
        return "", 204

    @app.get("/api/csrf")
    def api_csrf():
        # devolve o token csrf atual ou cria um novo se ainda não existir
        token = request.cookies.get("csrf_token") or secrets.token_urlsafe(32)
        response = jsonify({"csrf_token": token})
        response.set_cookie(
            "csrf_token",
            token,
            httponly=False,
            samesite="Strict",
            secure=False,
        )
        return response

    @app.get("/api/conteudo")
    def api_conteudo():
        # devolve todos os conteúdos oficiais (pontos, restaurantes, alojamentos, atividades) em JSON
        return jsonify(deps.obter_dados_projeto(origem="oficial"))

    @app.get("/api/recomendacoes")
    def api_recomendacoes():
        # devolve apenas os registos com origem 'malta' (sugestões aceites da comunidade)
        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            pontos = conn.execute(
                """
                SELECT id, nome, descricao, local, imagem, telefone, site_url, aviso, recomendado_por, lat, lon
                FROM pontos_interesse
                WHERE origem = 'malta'
                ORDER BY id DESC
                """
            ).fetchall()
            atividades = conn.execute(
                """
                SELECT id, nome, descricao, imagem, site_url, local, telefone, recomendado_por, lat, lon
                FROM atividades
                WHERE origem = 'malta'
                ORDER BY id DESC
                """
            ).fetchall()
            restaurantes = conn.execute(
                """
                SELECT id, nome, tipo, local, descricao, imagem, telefone, email, site_url, recomendado_por, lat, lon
                FROM restaurantes
                WHERE origem = 'malta'
                ORDER BY id DESC
                """
            ).fetchall()
            alojamentos = conn.execute(
                """
                SELECT id, nome, descricao, imagem, site_url, booking_url, local, telefone, recomendado_por, avaliacao_booking, lat, lon
                FROM alojamentos
                WHERE origem = 'malta'
                ORDER BY id DESC
                """
            ).fetchall()

        return jsonify(
            {
                "pontos": [dict(item) for item in pontos],
                "atividades": [dict(item) for item in atividades],
                "gastronomia": [dict(item) for item in restaurantes],
                "alojamentos": [dict(item) for item in alojamentos],
            }
        )

    @app.get("/api/restaurantes/telefones")
    def api_restaurantes_telefones():
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        forcar = str(request.args.get("forcar", "")).strip().lower() in ("1", "true", "sim", "yes")
        restaurantes = deps.ler_tabela_projeto("restaurantes")
        resultados = deps.buscar_telefones_restaurantes(restaurantes, forcar=forcar)
        return jsonify(resultados)

    @app.get("/api/admin/contactos")
    def api_admin_contactos():
        # devolve os últimos 100 contactos (requer token admin)
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            contactos = conn.execute(
                """
                SELECT id, nome, email, telefone, mensagem, estado, criado_em
                FROM contactos
                ORDER BY id DESC
                LIMIT 100
                """
            ).fetchall()

        return jsonify([dict(contacto) for contacto in contactos])

    @app.get("/api/admin/resumo")
    def api_admin_resumo():
        # devolve contagens globais (contactos, sugestões, logs) para o painel de resumo
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            resumo = {
                "contactos": conn.execute("SELECT COUNT(*) FROM contactos").fetchone()[0],
                "contactos_pendentes": conn.execute(
                    "SELECT COUNT(*) FROM contactos WHERE estado = 'pendente'"
                ).fetchone()[0],
                "sugestoes_pendentes": conn.execute(
                    "SELECT COUNT(*) FROM sugestoes_locais WHERE estado = 'pendente'"
                ).fetchone()[0],
                "sugestoes_aceites": conn.execute(
                    "SELECT COUNT(*) FROM sugestoes_locais WHERE estado = 'aceite'"
                ).fetchone()[0],
            }

        with sqlite3.connect(deps.DB_PATH) as conn:
            resumo["logs"] = conn.execute("SELECT COUNT(*) FROM api_logs").fetchone()[0]

        return jsonify(resumo)

    @app.get("/api/admin/contactos/exportar.xlsx")
    def api_admin_contactos_exportar():
        # gera e devolve um ficheiro Excel com todos os contactos marcados como resolvidos
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            contactos = conn.execute(
                """
                SELECT id, nome, email, telefone, mensagem, estado, criado_em
                FROM contactos
                WHERE estado = 'resolvido'
                ORDER BY id DESC
                """
            ).fetchall()

        colunas = ["id", "nome", "email", "telefone", "mensagem", "estado", "criado_em"]
        cabecalhos = ["ID", "Nome", "E-mail", "Telefone", "Mensagem", "Estado", "Data e hora"]

        def formatar_data_hora(valor: str) -> str:
            try:
                return datetime.fromisoformat(valor).strftime("%d/%m/%Y, %H:%M:%S")
            except (TypeError, ValueError):
                return valor or ""

        livro = Workbook()
        folha = livro.active
        folha.title = "Contactos resolvidos"

        folha.append(cabecalhos)
        for celula in folha[1]:
            celula.font = Font(bold=True)

        for contacto in contactos:
            linha = dict(contacto)
            linha["estado"] = "Resolvido"
            linha["criado_em"] = formatar_data_hora(linha.get("criado_em"))
            folha.append([linha.get(coluna, "") for coluna in colunas])

        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        for linha in folha.iter_rows():
            for celula in linha:
                celula.alignment = alinhamento_centro

        # ajusta a largura de cada coluna ao conteúdo mais longo (cabeçalho incluído)
        for indice in range(1, len(colunas) + 1):
            valores = [cabecalhos[indice - 1]] + [
                str(linha[indice - 1].value) if linha[indice - 1].value is not None else ""
                for linha in folha.iter_rows(min_row=2)
            ]
            maior_largura = max(len(valor) for valor in valores)
            folha.column_dimensions[get_column_letter(indice)].width = maior_largura + 2

        buffer = io.BytesIO()
        livro.save(buffer)
        buffer.seek(0)

        return Response(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=contactos-sesimbra.xlsx"},
        )

    @app.post("/api/admin/contactos/<int:contacto_id>/resolver")
    def api_admin_contacto_resolver(contacto_id: int):
        # marca um contacto como 'resolvido' no painel admin
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            cursor = conn.execute(
                "UPDATE contactos SET estado = 'resolvido' WHERE id = ?",
                (contacto_id,),
            )

        if cursor.rowcount == 0:
            return jsonify({"erro": "Contacto não encontrado."}), 404

        return jsonify({"estado": "ok", "contacto_id": contacto_id, "acao": "resolvido"})

    @app.post("/api/admin/contactos/<int:contacto_id>/apagar")
    def api_admin_contacto_apagar(contacto_id: int):
        # apaga permanentemente um contacto da base de dados
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            cursor = conn.execute("DELETE FROM contactos WHERE id = ?", (contacto_id,))

        if cursor.rowcount == 0:
            return jsonify({"erro": "Contacto não encontrado."}), 404

        return jsonify({"estado": "ok", "contacto_id": contacto_id, "acao": "apagado"})

    @app.get("/api/admin/sugestoes")
    def api_admin_sugestoes():
        # lista as sugestões da comunidade filtradas por estado (pendente, aceite, todas)
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        estado = str(request.args.get("estado", "pendente")).strip().lower()
        estados_validos = {"pendente", "aceite", "recusada", "todas"}
        if estado not in estados_validos:
            return jsonify({"erro": "Estado inválido."}), 400

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            if estado == "todas":
                sugestoes = conn.execute(
                    """
                    SELECT id, categoria, nome, descricao, morada, telefone, site, booking_url, email, foto_path, recomendado_por, pagina_origem, lat, lon, estado, criado_em
                    FROM sugestoes_locais
                    ORDER BY id DESC
                    LIMIT 100
                    """
                ).fetchall()
            else:
                sugestoes = conn.execute(
                    """
                    SELECT id, categoria, nome, descricao, morada, telefone, site, booking_url, email, foto_path, recomendado_por, pagina_origem, lat, lon, estado, criado_em
                    FROM sugestoes_locais
                    WHERE estado = ?
                    ORDER BY id DESC
                    LIMIT 100
                    """,
                    (estado,),
                ).fetchall()

        return jsonify([dict(sugestao) for sugestao in sugestoes])

    @app.post("/api/admin/sugestoes")
    def api_admin_criar_sugestao():
        # cria uma sugestão diretamente a partir do admin e publica-a de imediato
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        dados = request.form.to_dict() if request.form else (request.get_json(silent=True) or {})
        foto_path = ""
        foto = request.files.get("foto")
        if foto and foto.filename:
            foto_path, erro_foto = deps.guardar_imagem_sugestao(foto)
            if erro_foto:
                return jsonify({"erro": erro_foto}), 400

        dados["foto_path"] = foto_path
        sugestao, erros = deps.validar_sugestao_local(dados)

        if erros:
            caminho_foto = deps.BASE_DIR / foto_path
            if caminho_foto.exists() and caminho_foto.is_file():
                caminho_foto.unlink()
            return jsonify({"erro": " ".join(erros)}), 400

        if not sugestao.get("foto_path"):
            return jsonify({"erro": "Adiciona uma foto antes de publicar a sugestão."}), 400

        nova_id, _ = deps.guardar_sugestao_local(sugestao)

        sugestao_publicada = deps.publicar_sugestao_local(nova_id)
        if not sugestao_publicada:
            deps.apagar_sugestao_local(nova_id)
            return jsonify({"erro": "Não foi possível publicar a sugestão."}), 400

        return jsonify({"estado": "ok", "sugestao_id": nova_id, "acao": "aceite"}), 201

    @app.post("/api/admin/sugestoes/<int:sugestao_id>/<acao>")
    def api_admin_moderar_sugestao(sugestao_id: int, acao: str):
        # aceitar publica a sugestão nas tabelas oficiais; recusar apaga-a
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        if acao not in ("aceitar", "recusar"):
            return jsonify({"erro": "Ação inválida."}), 400

        if acao == "recusar":
            if not deps.apagar_sugestao_local(sugestao_id):
                return jsonify({"erro": "Sugestão não encontrada."}), 404
            return jsonify({"estado": "ok", "sugestao_id": sugestao_id, "acao": "recusada"})

        dados_admin = request.form.to_dict() if request.form else (request.get_json(silent=True) or {})
        foto_admin = request.files.get("foto")
        if foto_admin and foto_admin.filename:
            foto_path, erro_foto = deps.guardar_imagem_sugestao(foto_admin)
            if erro_foto:
                return jsonify({"erro": erro_foto}), 400
            dados_admin["foto_path"] = foto_path
        elif not dados_admin.get("foto_path"):
            with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
                linha_foto = conn.execute(
                    "SELECT foto_path FROM sugestoes_locais WHERE id = ?",
                    (sugestao_id,),
                ).fetchone()
            if linha_foto and not linha_foto[0]:
                return jsonify({"erro": "Adiciona uma foto antes de aceitar a sugestão."}), 400

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            linha_sugestao = conn.execute(
                """
                SELECT id, categoria, nome, descricao, morada, telefone, site, booking_url, email, foto_path, recomendado_por, pagina_origem, lat, lon, estado, criado_em
                FROM sugestoes_locais
                WHERE id = ?
                """,
                (sugestao_id,),
            ).fetchone()

        if not linha_sugestao:
            return jsonify({"erro": "Sugestão não encontrada."}), 404

        sugestao_para_validar = dict(linha_sugestao)
        for campo in ("categoria", "nome", "descricao", "morada", "telefone", "site", "booking_url", "email", "recomendado_por", "foto_path", "lat", "lon"):
            if campo in dados_admin:
                sugestao_para_validar[campo] = dados_admin.get(campo)

        _, erros_validacao = deps.validar_sugestao_local(sugestao_para_validar)
        if erros_validacao:
            foto_path = dados_admin.get("foto_path")
            if foto_path:
                caminho_foto = deps.BASE_DIR / foto_path
                if caminho_foto.exists() and caminho_foto.is_file():
                    caminho_foto.unlink()
            return jsonify({"erro": " ".join(erros_validacao)}), 400

        sugestao = deps.publicar_sugestao_local(sugestao_id, dados_admin=dados_admin)
        if not sugestao:
            foto_path = dados_admin.get("foto_path")
            if foto_path:
                caminho_foto = deps.BASE_DIR / foto_path
                if caminho_foto.exists() and caminho_foto.is_file():
                    caminho_foto.unlink()
            return jsonify({"erro": "Não foi possível publicar a sugestão."}), 400

        return jsonify({"estado": "ok", "sugestao_id": sugestao_id, "acao": "aceite"})

    @app.patch("/api/admin/sugestoes/<int:sugestao_id>/coordenadas")
    def api_admin_atualizar_coordenadas(sugestao_id: int):
        # atualiza as coordenadas de uma sugestão e, se já aceite, também na tabela de conteúdo
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        dados = request.get_json(silent=True) or {}
        lat_str = str(dados.get("lat", "") or "").strip()
        lon_str = str(dados.get("lon", "") or "").strip()

        lat = None
        lon = None
        if lat_str or lon_str:
            if not (lat_str and lon_str):
                return jsonify({"erro": "Preenche latitude e longitude em conjunto."}), 400
            try:
                lat = float(lat_str)
                lon = float(lon_str)
                if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                    raise ValueError()
            except ValueError:
                return jsonify({"erro": "Coordenadas inválidas."}), 400

            if not (38.30 <= lat <= 38.60 and -9.35 <= lon <= -8.80):
                return jsonify({"erro": "Esta localização não fica em Sesimbra."}), 400

        tabela_por_categoria = {
            "Ponto turistico": "pontos_interesse",
            "Atividade": "atividades",
            "Gastronomia": "restaurantes",
            "Alojamento": "alojamentos",
        }

        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            linha = conn.execute(
                "SELECT categoria, estado FROM sugestoes_locais WHERE id = ?",
                (sugestao_id,)
            ).fetchone()
            if not linha:
                return jsonify({"erro": "Sugestão não encontrada."}), 404

            conn.execute(
                "UPDATE sugestoes_locais SET lat = ?, lon = ? WHERE id = ?",
                (lat, lon, sugestao_id)
            )

            if linha["estado"] == "aceite":
                tabela = tabela_por_categoria.get(linha["categoria"])
                if tabela:
                    conn.execute(
                        f"UPDATE {tabela} SET lat = ?, lon = ? WHERE origem_sugestao_id = ?",
                        (lat, lon, sugestao_id)
                    )

        return jsonify({"estado": "ok"})

    @app.get("/api/admin/coordenadas")
    def api_admin_coordenadas():
        # geocodifica um nome de local (via Nominatim) e devolve latitude e longitude
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        consulta = str(request.args.get("q", "")).strip()
        if len(consulta) < 3:
            return jsonify({"erro": "Indica um nome ou localização com pelo menos 3 caracteres."}), 400

        coordenadas = deps.geocodificar_local(consulta)
        if not coordenadas:
            return jsonify({"erro": "Não foi possível encontrar coordenadas para essa localização."}), 404

        lat, lon = coordenadas
        return jsonify({"lat": lat, "lon": lon})

    @app.get("/api/admin/conteudo/<tabela>")
    def api_admin_listar_conteudo(tabela: str):
        # lista os registos oficiais de uma tabela para edição no painel admin
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401
        if tabela not in CAMPOS_EDITAVEIS_CONTEUDO:
            return jsonify({"erro": "Tabela inválida."}), 400
        campos = ["id", "nome", "local", "lat", "lon", "origem"] + [
            c for c in CAMPOS_EDITAVEIS_CONTEUDO[tabela]
            if c not in ("nome", "local", "lat", "lon", "ordem")
        ]
        campos_unicos = list(dict.fromkeys(campos + ["ordem"]))
        campos_str = ", ".join(campos_unicos)
        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            linhas = conn.execute(
                f"SELECT {campos_str} FROM {tabela} WHERE origem = 'oficial' ORDER BY ordem, id"
            ).fetchall()
        return jsonify([dict(l) for l in linhas])

    @app.patch("/api/admin/conteudo/<tabela>/<int:item_id>")
    def api_admin_editar_conteudo(tabela: str, item_id: int):
        # atualiza campos de um item de conteúdo oficial (só os campos em CAMPOS_EDITAVEIS_CONTEUDO)
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401
        if tabela not in CAMPOS_EDITAVEIS_CONTEUDO:
            return jsonify({"erro": "Tabela inválida."}), 400
        dados = request.form.to_dict() if request.form else (request.get_json(silent=True) or {})
        permitidos = CAMPOS_EDITAVEIS_CONTEUDO[tabela]
        atualizacoes = {}
        lat_val = None
        lon_val = None
        for campo in permitidos:
            if campo not in dados:
                continue
            valor = dados[campo]
            if campo in ("lat", "lon"):
                if valor == "" or valor is None:
                    atualizacoes[campo] = None
                else:
                    try:
                        f = float(str(valor))
                        if campo == "lat" and not (-90 <= f <= 90):
                            raise ValueError()
                        if campo == "lon" and not (-180 <= f <= 180):
                            raise ValueError()
                        atualizacoes[campo] = f
                        if campo == "lat":
                            lat_val = f
                        else:
                            lon_val = f
                    except ValueError:
                        return jsonify({"erro": f"Valor inválido para {campo}."}), 400
            elif campo == "ordem":
                try:
                    atualizacoes[campo] = int(str(valor)) if str(valor).strip() != "" else 0
                except ValueError:
                    return jsonify({"erro": "Ordem deve ser um número."}), 400
            elif campo == "links_json":
                try:
                    json.loads(valor)
                    atualizacoes[campo] = valor
                except (ValueError, TypeError):
                    atualizacoes[campo] = "[]"
            else:
                valor_str = str(valor).strip() if valor is not None else ""
                if campo == "telefone" and valor_str:
                    if not deps.telefone_internacional_valido(valor_str):
                        return jsonify({"erro": "Telemóvel inválido. Use o indicativo internacional (ex: +351 912 345 678)."}), 400
                atualizacoes[campo] = valor_str or None
        foto = request.files.get("foto")
        if foto and foto.filename:
            imagem_path, erro_foto = deps.guardar_imagem_sugestao(foto)
            if erro_foto:
                return jsonify({"erro": erro_foto}), 400
            atualizacoes["imagem"] = imagem_path
        if not atualizacoes:
            return jsonify({"erro": "Nenhum campo para atualizar."}), 400
        if lat_val is not None and lon_val is not None:
            if not (38.30 <= lat_val <= 38.60 and -9.35 <= lon_val <= -8.80):
                return jsonify({"erro": "Esta localização não fica em Sesimbra."}), 400
        set_clause = ", ".join(f"{c} = ?" for c in atualizacoes)
        valores = list(atualizacoes.values()) + [item_id]
        with sqlite3.connect(deps.CONTACTOS_DB_PATH) as conn:
            cursor = conn.execute(
                f"UPDATE {tabela} SET {set_clause} WHERE id = ?", valores
            )
            if cursor.rowcount == 0:
                return jsonify({"erro": "Item não encontrado."}), 404
        return jsonify({"estado": "ok"})

    @app.get("/api/logs")
    def api_logs():
        # devolve os últimos 20 registos de chamadas ao assistente IA para o painel admin
        if not deps.validar_admin_token():
            return jsonify({"erro": "Acesso não autorizado."}), 401

        with sqlite3.connect(deps.DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            linhas = conn.execute(
                """
                SELECT id, created_at, endpoint, modelo, pedido_json, resposta_json, prompt, resposta_text, estado
                FROM api_logs
                ORDER BY id DESC
                LIMIT 20
                """
            ).fetchall()

        logs = []
        for linha in linhas:
            item = dict(linha)
            item.update(deps.resumo_resposta_gemini(item.get("resposta_json", "{}")))
            logs.append(item)

        return jsonify(logs)

    @app.get("/")
    def index():
        # redireciona a raiz do site para a página principal sesimbra.html
        return send_from_directory(deps.FRONTEND_DIR, "sesimbra.html")

    @app.get("/<path:filename>")
    def ficheiros_site(filename: str):
        # serve html das páginas, assets e imagens; qualquer outro caminho devolve 404
        if filename.endswith(".html"):
            return send_from_directory(deps.FRONTEND_DIR, filename)
        if filename.startswith(("assets/", "img/")):
            return send_from_directory(deps.BASE_DIR, filename)

        return send_from_directory(deps.FRONTEND_DIR, "404.html"), 404

    @app.errorhandler(404)
    def pagina_nao_encontrada(_erro):
        return send_from_directory(deps.FRONTEND_DIR, "404.html"), 404
